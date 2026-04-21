"""Generate NOAH-Prototype-Sprint-Plan.xlsx mirroring the format of NOAH-Sprint-Plan.xlsx.

Covers the prototype build: Option 3 (UC-07 → UC-10) + Narrative Agent (UC-18 + UC-20) + Canned/Pages + Packaging.

Sheets:
  1. Scenarios    — prototype build scenarios → PUCs → NOAH UCs → personas → sprints → acceptance → priority → claude hours
  2. By Sprint    — PUCs grouped by prototype sprint (PS-01..PS-Final) with subtasks and hours
  3. By Persona   — PUCs grouped by who drives/validates
  4. Traceability — PSC → PUC → NOAH UC → Sprint → Files → Hours
  5. Tech Inventory — files/modules created per PUC
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# ---------- Data ----------

SCENARIOS = [
    # PSC-01 Foundation
    ("PSC-01", "Foundation — Dual-Mode Scaffold", "PUC-01", "Vite + React + TS scaffold with dual-mode build",
     "Dev", "PS-01", "—",
     "VITE_MODE=live|canned drives both agent adapter and theme. Scaffold runs on fresh Windows laptop with pnpm.",
     "P0", "1-2h"),
    ("PSC-01", "Foundation — Dual-Mode Scaffold", "PUC-02", "Ollama client with JSON mode + health check",
     "Dev", "PS-01", "—",
     "chatJSON() returns validated JSON or throws OllamaError; checkHealth() pings /api/tags. Unit tests green.",
     "P0", "2-3h"),
    ("PSC-01", "Foundation — Dual-Mode Scaffold", "PUC-03", "Document ingest (PDF + DOCX) with page spans + SHA-256",
     "Dev", "PS-01", "—",
     "Ingests .pdf (pdfjs-dist) and .docx (mammoth). Returns fullText + pages + hash. Errors on scanned/empty.",
     "P0", "2-3h"),
    ("PSC-01", "Foundation — Dual-Mode Scaffold", "PUC-04", "Port close-cycle simulation to React",
     "Nike (viewer)", "PS-01", "UC-13,14,15",
     "4 phases (Pre-Close→Execute→Validate→Gate), 6 entities, event log, journey map. Matches NikeR2R-v4-Compare.html fidelity.",
     "P0", "3-4h"),
    ("PSC-01", "Foundation — Dual-Mode Scaffold", "PUC-05", "Port Copilot chat panel (scripted)",
     "Nike (viewer)", "PS-01", "UC-25",
     "Canned topic routing with 5+ keyword matches: close status, exceptions, contracts, accruals, default.",
     "P1", "1h"),

    # PSC-02 UC-07
    ("PSC-02", "Contract Attribute Extraction (UC-07)", "PUC-06", "27-attribute zod schema",
     "Dev", "PS-02", "UC-07",
     "Schema enumerates 27 attribute names; each field = {value, confidence, source_page}. Validation rejects malformed.",
     "P0", "0.5h"),
    ("PSC-02", "Contract Attribute Extraction (UC-07)", "PUC-07", "Extractor agent calling Qwen",
     "Keven (demo)", "PS-02", "UC-07",
     "Real Qwen JSON-mode call returns validated ContractAttributes. Retries once with correction prompt on bad JSON.",
     "P0", "2-3h"),
    ("PSC-02", "Contract Attribute Extraction (UC-07)", "PUC-08", "Agent adapter interface + LiveAgent",
     "Dev", "PS-02", "—",
     "Agent interface has extract/risk/tech/accrual/narrative methods. LiveAgent emits step events for UI strip.",
     "P0", "1h"),
    ("PSC-02", "Contract Attribute Extraction (UC-07)", "PUC-09", "Contract queue screen",
     "Keven (demo)", "PS-02", "UC-08",
     "Drag-drop upload (.pdf/.docx). Seed contracts + uploaded list visible. Sorts by risk.",
     "P0", "2h"),
    ("PSC-02", "Contract Attribute Extraction (UC-07)", "PUC-10", "Contract review screen",
     "Nike (viewer)", "PS-02", "UC-07",
     "Document viewer + 27-attribute checklist + agent activity strip + confidence badges (green/amber/red).",
     "P0", "2-3h"),
    ("PSC-02", "Contract Attribute Extraction (UC-07)", "PUC-11", "Ollama pre-flight guard",
     "Dev", "PS-02", "—",
     "Modal blocks /contracts and /narrative routes if Ollama is unhealthy; Retry re-checks. Close Cockpit still loads.",
     "P1", "0.5-1h"),

    # PSC-03 UC-08 + UC-09
    ("PSC-03", "Risk + Technical Accounting (UC-08, UC-09)", "PUC-12", "Risk scoring (rules + LLM signal)",
     "Dev", "PS-03", "UC-08",
     "scoreRules + scoreLLM; category High/Med/Low; reasons[].",
     "P0", "1-2h"),
    ("PSC-03", "Risk + Technical Accounting (UC-08, UC-09)", "PUC-13", "Technical accounting classifier",
     "Dev", "PS-03", "UC-09",
     "Detects ASC 840/842 leases + ASC 815 derivatives. Returns expense recognition method + requires_senior_review flag.",
     "P0", "2-3h"),
    ("PSC-03", "Risk + Technical Accounting (UC-08, UC-09)", "PUC-14", "Risk + flags UI",
     "Nike (viewer)", "PS-03", "UC-08,09",
     "Risk gauge + category badge + reasons panel. ASC tags with mandatory senior review banner when flagged.",
     "P0", "1-2h"),

    # PSC-04 UC-10
    ("PSC-04", "Accrual & JE (UC-10)", "PUC-15", "Accrual inputs extractor (strings/dates only)",
     "Dev", "PS-04", "UC-10",
     "AccrualInputs type has NO numeric fields. LLM returns strings/dates; missing[] populated on low-confidence gaps.",
     "P0", "1-2h"),
    ("PSC-04", "Accrual & JE (UC-10)", "PUC-16", "Deterministic accrual math",
     "Dev", "PS-04", "UC-10",
     "Pure TS: straight-line pro-ration, GR/IR net (billedToDate subtracted), zero-period handling. TDD with 3+ cases.",
     "P0", "2-3h"),
    ("PSC-04", "Accrual & JE (UC-10)", "PUC-17", "JE builder",
     "Dev", "PS-04", "UC-10",
     "Two-line entry (DR Expense / CR Accrued Liability). Debits = credits. Reversal = 1st of next month.",
     "P0", "1h"),
    ("PSC-04", "Accrual & JE (UC-10)", "PUC-18", "Accrual agent orchestration",
     "Dev", "PS-04", "UC-10",
     "Extract → parse → math → build JE. Emits step events. Throws with field list on missing inputs.",
     "P0", "1h"),
    ("PSC-04", "Accrual & JE (UC-10)", "PUC-19", "Accrual proposal screen",
     "Nike (viewer)", "PS-04", "UC-10",
     "JE card (T-account), calc detail panel, clause traceability links, Approve button → close-store audit event.",
     "P0", "2-3h"),

    # PSC-05 Narrative (NEW)
    ("PSC-05", "Narrative Agent — Variance & Exec Summary (UC-18, UC-20)", "PUC-20", "Seed P&L dataset",
     "Dev", "PS-05", "UC-18,20",
     "12-15 line items: current vs prior period, variance $ and %, drivers, entity split (NA/EMEA/GC/APLA). Rich enough for Qwen to ground.",
     "P0", "1-2h"),
    ("PSC-05", "Narrative Agent — Variance & Exec Summary (UC-18, UC-20)", "PUC-21", "Variance commentary agent",
     "Dev", "PS-05", "UC-18",
     "Qwen generates 2-3 sentence prose per line item grounded ONLY in supplied numbers. Key drivers + risk flags + confidence.",
     "P0", "1-2h"),
    ("PSC-05", "Narrative Agent — Variance & Exec Summary (UC-18, UC-20)", "PUC-22", "Executive summary agent",
     "Dev", "PS-05", "UC-20",
     "Takes close metrics + top variances + risks → headline + key highlights + risks + recommendation. Board-ready tone.",
     "P0", "1h"),
    ("PSC-05", "Narrative Agent — Variance & Exec Summary (UC-18, UC-20)", "PUC-23", "Adapter: narrative methods",
     "Dev", "PS-05", "—",
     "Extend Agent interface with generateVarianceCommentary + generateExecutiveSummary. LiveAgent emits events.",
     "P0", "0.5h"),
    ("PSC-05", "Narrative Agent — Variance & Exec Summary (UC-18, UC-20)", "PUC-24", "Variance commentary screen",
     "Nike (viewer)", "PS-05", "UC-18",
     "Narrative route. P&L grid + per-row Generate button + side panel showing streaming prose, drivers, risks. Copy-to-clipboard.",
     "P0", "2h"),
    ("PSC-05", "Narrative Agent — Variance & Exec Summary (UC-18, UC-20)", "PUC-25", "Executive summary tab + Cockpit integration",
     "Nike (viewer)", "PS-05", "UC-20",
     "Narrative screen Exec Summary tab. Close Cockpit shows Generate Close Narrative button at Gate phase; navigates to /narrative?tab=exec&autorun.",
     "P0", "1-2h"),

    # PSC-06 Canned + Pages (was PSC-05)
    ("PSC-06", "Canned Mode + GitHub Pages", "PUC-26", "CannedAgent adapter (fixture replay)",
     "Dev", "PS-06", "—",
     "Loads fixtures/<hash>.json for contracts, plus narrative fixtures keyed by line-item ID. Replays with 200-800ms step delays.",
     "P0", "1-2h"),
    ("PSC-06", "Canned Mode + GitHub Pages", "PUC-27", "Fixture generator script",
     "Dev", "PS-06", "—",
     "pnpm generate-fixtures walks samples/acme/*.{pdf,docx} and seed-pnl, runs LiveAgent chain, writes fixtures/*.json.",
     "P0", "1h"),
    ("PSC-06", "Canned Mode + GitHub Pages", "PUC-28", "Sample contracts in place (5 supplied)",
     "Keven (demo)", "PS-06", "—",
     "Contract_1_Advertising_Campaign, Contract_2_Professional_Services_Outsourcing, Contract_3_Insurance_MultiYear, Contract_4_Construction_Retail_Remodel, Contract_5_AWS_Enterprise — already in samples/acme/.",
     "P0", "0.5h"),
    ("PSC-06", "Canned Mode + GitHub Pages", "PUC-29", "GitHub Pages deploy workflow",
     "Dev", "PS-06", "—",
     "On push to main: pnpm build:pages → deploy to gh-pages. Public URL loads Acme demo.",
     "P0", "1h"),

    # PSC-07 Packaging & QA (was PSC-06)
    ("PSC-07", "Packaging & QA", "PUC-30", "Full README + demo script",
     "Keven (demo)", "PS-Final", "—",
     "Both setup paths. 60-second demo script covering all 3 agent flows (contract, JE, narrative). Troubleshooting.",
     "P0", "2h"),
    ("PSC-07", "Packaging & QA", "PUC-31", "Prompt regression harness",
     "Dev", "PS-Final", "—",
     "pnpm test:prompts runs LiveAgent against 5 Acme contracts + narrative golden set; diffs against expected.",
     "P1", "1h"),
    ("PSC-07", "Packaging & QA", "PUC-32", "Playwright E2E (canned mode)",
     "Dev", "PS-Final", "—",
     "Upload → extract → review → propose → approve flow passes. Narrative tab generates commentary. Deterministic assertions.",
     "P1", "1-2h"),
    ("PSC-07", "Packaging & QA", "PUC-33", "Pre-commit hook blocking Deloitte docs",
     "Dev", "PS-Final", "—",
     "husky pre-commit rejects any staged file under samples/user/ (except .gitkeep). Blocks .pdf and .docx.",
     "P0", "0.5h"),
    ("PSC-07", "Packaging & QA", "PUC-34", "Dress rehearsal on clean machine",
     "Keven (demo)", "PS-Final", "—",
     "Fresh Windows laptop: clone → setup < 5 min → run full demo both modes including narrative. Polish rough edges.",
     "P0", "1-2h"),
]

FILES_BY_PUC = {
    "PUC-01": ("package.json, vite.config.ts, src/main.tsx, src/App.tsx, tailwind.config.js", "—", "PS-01"),
    "PUC-02": ("src/agents/ollama-client.ts", "src/agents/ollama-client.test.ts", "PS-01"),
    "PUC-03": ("src/lib/ingest.ts (pdfjs-dist + mammoth)", "src/lib/ingest.test.ts", "PS-01"),
    "PUC-04": ("src/screens/CloseCockpit.tsx, src/store/closeStore.ts, src/components/{PhaseGrid,EntityList,EventLog}.tsx", "manual smoke", "PS-01"),
    "PUC-05": ("src/screens/CopilotPanel.tsx, src/data/copilot-canned-answers.ts", "manual smoke", "PS-01"),
    "PUC-06": ("src/agents/contract-schema.ts", "src/agents/contract-schema.test.ts", "PS-02"),
    "PUC-07": ("src/agents/extractor.ts", "src/agents/extractor.test.ts", "PS-02"),
    "PUC-08": ("src/adapters/{agent-interface,live-agent,index}.ts", "—", "PS-02"),
    "PUC-09": ("src/screens/ContractQueue.tsx, src/store/contractStore.ts, src/data/seed-contracts.ts", "manual smoke", "PS-02"),
    "PUC-10": ("src/screens/ContractReview.tsx, src/components/{AgentActivityStrip,ConfidenceBadge,AttributeChecklist,DocViewer}.tsx", "manual E2E", "PS-02"),
    "PUC-11": ("src/components/OllamaGuard.tsx", "manual failure test", "PS-02"),
    "PUC-12": ("src/agents/risk.ts", "src/agents/risk.test.ts", "PS-03"),
    "PUC-13": ("src/agents/tech-accounting.ts", "src/agents/tech-accounting.test.ts", "PS-03"),
    "PUC-14": ("src/components/{RiskPanel,TechAccountingFlags}.tsx, modify ContractReview.tsx", "manual smoke", "PS-03"),
    "PUC-15": ("src/agents/accrual-inputs.ts", "src/agents/accrual-inputs.test.ts", "PS-04"),
    "PUC-16": ("src/lib/accrual-math.ts", "src/lib/accrual-math.test.ts (TDD)", "PS-04"),
    "PUC-17": ("src/lib/je-builder.ts", "src/lib/je-builder.test.ts", "PS-04"),
    "PUC-18": ("src/agents/accrual.ts", "—", "PS-04"),
    "PUC-19": ("src/screens/AccrualProposal.tsx, src/components/{JECard,CalcDetailPanel}.tsx", "manual E2E", "PS-04"),
    "PUC-20": ("src/data/seed-pnl.ts", "—", "PS-05"),
    "PUC-21": ("src/agents/narrative.ts (generateVarianceCommentary)", "src/agents/narrative.test.ts", "PS-05"),
    "PUC-22": ("src/agents/narrative.ts (generateExecutiveSummary)", "src/agents/narrative.test.ts", "PS-05"),
    "PUC-23": ("modify src/adapters/{agent-interface,live-agent}.ts", "—", "PS-05"),
    "PUC-24": ("src/screens/Narrative.tsx, src/components/{VarianceTable,CommentaryPanel}.tsx", "manual E2E", "PS-05"),
    "PUC-25": ("src/components/ExecSummaryCard.tsx, modify CloseCockpit.tsx + Narrative.tsx", "manual smoke", "PS-05"),
    "PUC-26": ("src/adapters/canned-agent.ts", "src/adapters/canned-agent.test.ts", "PS-06"),
    "PUC-27": ("src/scripts/generate-fixtures.ts", "—", "PS-06"),
    "PUC-28": ("samples/acme/Contract_{1..5}_*.docx (supplied), samples/README.md", "manual review", "PS-06"),
    "PUC-29": (".github/workflows/pages.yml", "CI run verification", "PS-06"),
    "PUC-30": ("README.md", "—", "PS-Final"),
    "PUC-31": ("src/scripts/prompt-regression.ts, tests/golden/*.expected.json", "script passes", "PS-Final"),
    "PUC-32": ("playwright.config.ts, tests/e2e/contract-flow.spec.ts, tests/e2e/narrative.spec.ts", "playwright green", "PS-Final"),
    "PUC-33": (".husky/pre-commit", "manual: attempt to commit contract in samples/user → rejected", "PS-Final"),
    "PUC-34": ("—", "dress rehearsal checklist", "PS-Final"),
}

SPRINT_DESCRIPTIONS = {
    "PS-01": ("Foundation", "Scaffold + Ollama client + PDF/DOCX ingest + port scripted simulations"),
    "PS-02": ("UC-07 Contract Extraction", "Real Qwen extraction of 27 attributes into contract review UI"),
    "PS-03": ("UC-08 + UC-09 Risk/Tech-Acct", "Risk scoring and ASC 840/842/815 classification"),
    "PS-04": ("UC-10 Accrual + JE", "Extract → deterministic math → proposed JE with approval"),
    "PS-05": ("UC-18 + UC-20 Narrative Agent", "Variance commentary + executive close narrative (3rd real-AI agent)"),
    "PS-06": ("Canned Mode + Pages", "Fixture-backed Acme theme for public GitHub Pages deployment"),
    "PS-Final": ("Packaging & QA", "Docs, regression tests, E2E, pre-commit guard, dress rehearsal"),
}

# ---------- Styling ----------

THIN = Side(style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="111111")
HEADER_FONT = Font(color="C8FF00", bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14, color="111111")
SUBTITLE_FONT = Font(italic=True, size=10, color="555555")
SECTION_FILL = PatternFill("solid", fgColor="F0F0F0")
SECTION_FONT = Font(bold=True, size=11, color="111111")
NARR_FILL = PatternFill("solid", fgColor="FFF9E5")   # highlight for new narrative scenario
P0_FILL = PatternFill("solid", fgColor="FFE5E5")
P1_FILL = PatternFill("solid", fgColor="FFF5E5")
P2_FILL = PatternFill("solid", fgColor="F0F0F0")
WRAP = Alignment(wrap_text=True, vertical="top")


def style_header_row(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
        cell.border = BORDER


def priority_fill(priority):
    return P0_FILL if priority.startswith("P0") else P1_FILL if priority.startswith("P1") else P2_FILL


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ---------- Build workbook ----------

wb = Workbook()

# ===== Sheet 1: Scenarios =====
ws = wb.active
ws.title = "Scenarios"

ws.cell(row=1, column=1, value="NOAH Prototype — Scenario & Use Case Matrix").font = TITLE_FONT
ws.cell(row=2, column=1, value="Prototype Build Plan — Option 3 (UC-07 → UC-10) + Narrative (UC-18, UC-20) + Canned/Pages + Packaging").font = SUBTITLE_FONT
ws.merge_cells("A1:J1")
ws.merge_cells("A2:J2")

headers = ["PSC #", "Scenario", "PUC #", "Prototype Use Case",
           "Persona", "Sprint", "NOAH UC(s)", "Acceptance Criteria", "Priority", "Claude Time"]
for i, h in enumerate(headers, 1):
    ws.cell(row=4, column=i, value=h)
style_header_row(ws, 4, len(headers))

row = 5
prev_psc = None
for psc, scenario, puc, uc_title, persona, sprint, noah_uc, accept, priority, hours in SCENARIOS:
    merge_cell = psc != prev_psc
    cells = [psc, scenario, puc, uc_title, persona, sprint, noah_uc, accept, priority, hours]
    is_narr = psc == "PSC-05"
    for i, v in enumerate(cells, 1):
        c = ws.cell(row=row, column=i, value=v)
        c.alignment = WRAP
        c.border = BORDER
        if is_narr:
            c.fill = NARR_FILL
        if i == 9:
            c.fill = priority_fill(priority)
    if not merge_cell:
        ws.cell(row=row, column=1, value=None)
        ws.cell(row=row, column=2, value=None)
    row += 1
    prev_psc = psc

set_widths(ws, [9, 34, 9, 34, 14, 10, 14, 60, 10, 12])
ws.row_dimensions[4].height = 32
ws.freeze_panes = "A5"

# ===== Sheet 2: By Sprint =====
ws2 = wb.create_sheet("By Sprint")
ws2.cell(row=1, column=1, value="Prototype Sprints — Tasks, Files, Hours").font = TITLE_FONT
ws2.merge_cells("A1:G1")

headers2 = ["Sprint", "PUC #", "Task", "Persona", "Files (create/modify)", "Tests", "Claude Time"]
for i, h in enumerate(headers2, 1):
    ws2.cell(row=3, column=i, value=h)
style_header_row(ws2, 3, len(headers2))

row = 4
by_sprint = {}
for psc, scenario, puc, uc_title, persona, sprint, noah_uc, accept, priority, hours in SCENARIOS:
    by_sprint.setdefault(sprint, []).append((puc, uc_title, persona, hours))

for sprint_key, (sprint_name, sprint_desc) in SPRINT_DESCRIPTIONS.items():
    fill = NARR_FILL if sprint_key == "PS-05" else SECTION_FILL
    c = ws2.cell(row=row, column=1, value=f"{sprint_key} — {sprint_name}")
    c.font = SECTION_FONT
    c.fill = fill
    c2 = ws2.cell(row=row, column=2, value=sprint_desc)
    c2.fill = fill
    c2.font = Font(italic=True, size=10)
    for col in range(3, 8):
        ws2.cell(row=row, column=col).fill = fill
    ws2.merge_cells(start_row=row, start_column=2, end_row=row, end_column=7)
    row += 1
    for puc, uc_title, persona, hours in by_sprint.get(sprint_key, []):
        files, tests, _ = FILES_BY_PUC.get(puc, ("—", "—", sprint_key))
        cells = [sprint_key, puc, uc_title, persona, files, tests, hours]
        for i, v in enumerate(cells, 1):
            c = ws2.cell(row=row, column=i, value=v)
            c.alignment = WRAP
            c.border = BORDER
            if sprint_key == "PS-05":
                c.fill = NARR_FILL
        row += 1

# Totals row
total_min, total_max = 0, 0
for _, _, _, _, _, _, _, _, _, hrs in SCENARIOS:
    h = hrs.replace("h", "").replace(" ", "")
    if "-" in h:
        lo, hi = h.split("-")
        total_min += float(lo); total_max += float(hi)
    else:
        total_min += float(h); total_max += float(h)
c = ws2.cell(row=row + 1, column=1, value="TOTAL")
c.font = Font(bold=True)
ws2.cell(row=row + 1, column=3, value="Estimated build time across all sprints").font = Font(italic=True)
t = ws2.cell(row=row + 1, column=7, value=f"{total_min:.0f}-{total_max:.0f}h")
t.font = Font(bold=True)

set_widths(ws2, [10, 9, 36, 14, 60, 28, 12])
ws2.row_dimensions[3].height = 30
ws2.freeze_panes = "A4"

# ===== Sheet 3: By Persona =====
ws3 = wb.create_sheet("By Persona")
ws3.cell(row=1, column=1, value="Prototype Build — Responsibilities by Persona").font = TITLE_FONT
ws3.merge_cells("A1:E1")

headers3 = ["PUC #", "Task", "Sprint", "Acceptance (summary)", "Claude Time"]
personas = sorted({r[4] for r in SCENARIOS})

row = 3
for persona in personas:
    c = ws3.cell(row=row, column=1, value=persona)
    c.font = SECTION_FONT
    c.fill = SECTION_FILL
    ws3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    row += 1
    for i, h in enumerate(headers3, 1):
        ws3.cell(row=row, column=i, value=h)
    style_header_row(ws3, row, len(headers3))
    row += 1
    for psc, scenario, puc, uc_title, pers, sprint, noah_uc, accept, priority, hours in SCENARIOS:
        if pers != persona:
            continue
        summary = accept if len(accept) < 120 else accept[:117] + "..."
        cells = [puc, uc_title, sprint, summary, hours]
        for i, v in enumerate(cells, 1):
            c = ws3.cell(row=row, column=i, value=v)
            c.alignment = WRAP
            c.border = BORDER
        row += 1
    row += 1

set_widths(ws3, [9, 36, 10, 60, 12])
ws3.freeze_panes = "A4"

# ===== Sheet 4: Traceability =====
ws4 = wb.create_sheet("Traceability")
ws4.cell(row=1, column=1, value="Traceability: Prototype Scenario → PUC → NOAH UC → Sprint → Files → Hours").font = TITLE_FONT
ws4.merge_cells("A1:G1")

headers4 = ["PSC", "PUC #", "Title", "NOAH UC", "Sprint", "Files", "Claude Time"]
for i, h in enumerate(headers4, 1):
    ws4.cell(row=3, column=i, value=h)
style_header_row(ws4, 3, len(headers4))

row = 4
for psc, scenario, puc, uc_title, persona, sprint, noah_uc, accept, priority, hours in SCENARIOS:
    files, _, _ = FILES_BY_PUC.get(puc, ("—", "—", sprint))
    cells = [psc, puc, uc_title, noah_uc, sprint, files, hours]
    for i, v in enumerate(cells, 1):
        c = ws4.cell(row=row, column=i, value=v)
        c.alignment = WRAP
        c.border = BORDER
        if psc == "PSC-05":
            c.fill = NARR_FILL
    row += 1

set_widths(ws4, [9, 9, 36, 14, 10, 55, 12])
ws4.freeze_panes = "A4"

# ===== Sheet 5: Tech Inventory =====
ws5 = wb.create_sheet("Tech Inventory")
ws5.cell(row=1, column=1, value="Files & Modules Created per PUC").font = TITLE_FONT
ws5.merge_cells("A1:D1")

headers5 = ["PUC #", "Files Created/Modified", "Tests", "Sprint"]
for i, h in enumerate(headers5, 1):
    ws5.cell(row=3, column=i, value=h)
style_header_row(ws5, 3, len(headers5))

row = 4
for puc, (files, tests, sprint) in FILES_BY_PUC.items():
    cells = [puc, files, tests, sprint]
    for i, v in enumerate(cells, 1):
        c = ws5.cell(row=row, column=i, value=v)
        c.alignment = WRAP
        c.border = BORDER
        if sprint == "PS-05":
            c.fill = NARR_FILL
    row += 1

set_widths(ws5, [9, 70, 32, 10])
ws5.freeze_panes = "A4"

# Save
OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "NOAH-Prototype-Sprint-Plan.xlsx")
wb.save(OUT)
print(f"Saved: {OUT}")
print(f"Total Claude time across {len(SCENARIOS)} PUCs: {total_min:.0f}-{total_max:.0f}h")
